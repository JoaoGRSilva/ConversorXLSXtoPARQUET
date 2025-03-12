import pandas as pd
import os
import time
import pyarrow as pa
import pyarrow.parquet as pq
from tqdm.auto import tqdm
import openpyxl
import gc  # Importar o coletor de lixo

def xlsx_para_parquet(file_path, parquet_file):
    print(f"Iniciando conversão de: {os.path.basename(file_path)}")
    
    try:
        # Configurações otimizadas
        batch_size = 25000  # Ajuste conforme necessário
        compression = 'snappy'  # Ou 'gzip', 'brotli' se desejar maior compressão
        
        # Obter informações sobre o arquivo Excel
        wb = openpyxl.load_workbook(file_path, read_only=True)
        sheet_name = wb.sheetnames[0]
        ws = wb[sheet_name]
        
        # Calcular o número total de linhas (excluindo o cabeçalho)
        total_rows = ws.max_row - 1
        
        # Inicializar a barra de progresso
        with tqdm(total=total_rows, desc="Lendo Excel", unit="rows") as pbar:
            # Inicializar lista para armazenar os DataFrames de cada lote
            dfs = []
            
            # Iterar sobre as linhas em lotes
            for i in range(2, ws.max_row + 1, batch_size):
                min_row = i
                max_row = min(i + batch_size - 1, ws.max_row)
                
                # Ler o lote atual para um DataFrame
                batch_data = ws.iter_rows(min_row=min_row, max_row=max_row, values_only=True)
                df = pd.DataFrame(batch_data, columns=[cell.value for cell in ws[1]])
                dfs.append(df)
                
                # Atualizar a barra de progresso
                pbar.update(len(df))
                
                # Limpar memória (descartar DataFrame temporário)
                del df
                gc.collect()
            
            # Concatenar todos os DataFrames em um único DataFrame
            df = pd.concat(dfs, ignore_index=True)
            
            # Converter para parquet
            print(f"\nConvertendo para formato Parquet...")
            table = pa.Table.from_pandas(df)
            pq.write_table(table, parquet_file, compression=compression)
            
            file_size_mb = os.path.getsize(parquet_file) / (1024 * 1024)
            
            print(f"\n✅ Conversão concluída com sucesso!")
            print(f"   Arquivo: {os.path.basename(parquet_file)}")
            print(f"   Tamanho: {file_size_mb:.2f} MB")
            print(f"   Linhas processadas: {len(df):,}")
        
        wb.close()
        
    except Exception as e:
        print(f"\n❌ Erro durante a conversão: {str(e)}")
        print("Não foi possível converter o arquivo.")

if __name__ == "__main__":
    # Verificar diretório atual e encontrar arquivos Excel
    directory = os.getcwd()
    files = os.listdir(directory)
    xlsx_files = [file for file in files if file.endswith('.xlsx') or file.endswith('.xls')]
    
    if xlsx_files:
        file_path = os.path.join(directory, xlsx_files[0])
        parquet_file = 'dados.parquet'
        
        print("="*60)
        print(f"🔄 CONVERSOR XLSX PARA PARQUET")
        print("="*60)
        print(f"📄 Arquivo encontrado: {xlsx_files[0]}")
        print(f"💾 Destino: {parquet_file}")
        print("-"*60)
        
        # Iniciar conversão
        xlsx_para_parquet(file_path, parquet_file)
    else:
        print("❌ Nenhum arquivo Excel (.xlsx ou .xls) encontrado no diretório atual.")